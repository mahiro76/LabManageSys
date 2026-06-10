"""实验室考勤与权限管理系统（UI 模块）。

包含 Application 主类和所有 Tab 页面类。
通过 `from ui import Application; Application().Run()` 启动。
"""

from __future__ import annotations

import csv
import os
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Callable, Dict, List, Optional, Tuple
import tkinter as tk

import mysql.connector
from openpyxl import load_workbook

import matplotlib

matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# ── matplotlib 中文字体配置 ──
def _SetupChineseFont() -> None:
    """为 matplotlib 配置支持中文的字体，避免图表中文显示为方框。"""
    import matplotlib.font_manager as fm
    # Windows 常见中文字体候选列表
    ChineseFontCandidates = [
        "Microsoft YaHei",           # 微软雅黑
        "SimHei",                    # 黑体
        "DengXian",                  # 等线
        "Microsoft YaHei UI",
        "SimSun",                    # 宋体
        "FangSong",                  # 仿宋
        "KaiTi",                     # 楷体
    ]
    AvailableFonts = [f.name for f in fm.fontManager.ttflist]
    for FontName in ChineseFontCandidates:
        if FontName in AvailableFonts:
            matplotlib.rcParams["font.family"] = FontName
            matplotlib.rcParams["font.sans-serif"] = [FontName]
            matplotlib.rcParams["axes.unicode_minus"] = False
            return
    # 若均未找到，尝试通过字体文件路径加载
    try:
        import os
        FontDir = r"C:\Windows\Fonts"
        FontFiles = ["msyh.ttc", "simhei.ttf", "deng.ttf", "simsun.ttc"]
        for Fname in FontFiles:
            Fpath = os.path.join(FontDir, Fname)
            if os.path.exists(Fpath):
                fm.fontManager.addfont(Fpath)
                FontName = fm.FontProperties(fname=Fpath).get_name()
                matplotlib.rcParams["font.family"] = FontName
                matplotlib.rcParams["font.sans-serif"] = [FontName]
                matplotlib.rcParams["axes.unicode_minus"] = False
                return
    except Exception:
        pass
    print("[WARN] 未找到中文字体，图表中文可能显示为方框")

_SetupChineseFont()

from db import MySqlManager
from models import (
    DatabaseConfig, LoginUser,
    ROLE_ADMIN, ROLE_STAFF, ROLE_STUDENT,
    ATTENDANCE_STATUSES, GENDER_OPTIONS,
    DEFAULT_DB_HOST, DEFAULT_DB_PORT, DEFAULT_DB_USER,
    DEFAULT_DB_NAME, DEFAULT_DB_PASSWORD,
    IsValidPhone, IsValidAge, IsValidDate, IsValidTime,
    NormalizeText, BuildNowText,
)

# 尝试导入本地 MySQL 服务管理（可选依赖）
try:
    from mysql_service import start_mysql, stop_mysql, initialize_mysql
except Exception:
    start_mysql = None
    stop_mysql = None
    initialize_mysql = None


# ========================================================================
#  工具函数
# ========================================================================

def _GetDefaultDbConfig() -> DatabaseConfig:
    """从 dbconfig.json 读取数据库配置，不存在则用默认值生成。"""
    import json
    ConfigPath = os.path.join(os.path.dirname(__file__), "dbconfig.json")
    if os.path.exists(ConfigPath):
        try:
            with open(ConfigPath, "r", encoding="utf-8") as F:
                Data = json.load(F)
            return DatabaseConfig(
                Host=Data.get("host", DEFAULT_DB_HOST),
                Port=int(Data.get("port", DEFAULT_DB_PORT)),
                User=Data.get("user", DEFAULT_DB_USER),
                Password=Data.get("password", DEFAULT_DB_PASSWORD),
                Database=Data.get("database", DEFAULT_DB_NAME),
            )
        except (json.JSONDecodeError, KeyError, ValueError):
            pass
    # 不存在或解析失败，用默认值生成
    Config = DatabaseConfig(
        Host=DEFAULT_DB_HOST,
        Port=int(DEFAULT_DB_PORT),
        User=DEFAULT_DB_USER,
        Password=DEFAULT_DB_PASSWORD,
        Database=DEFAULT_DB_NAME,
    )
    _SaveDbConfig(Config)
    return Config


def _SaveDbConfig(Config: DatabaseConfig) -> None:
    """将数据库配置保存到 dbconfig.json。"""
    import json
    ConfigPath = os.path.join(os.path.dirname(__file__), "dbconfig.json")
    with open(ConfigPath, "w", encoding="utf-8") as F:
        json.dump({
            "host": Config.Host,
            "port": Config.Port,
            "user": Config.User,
            "password": Config.Password,
            "database": Config.Database,
        }, F, ensure_ascii=False, indent=2)


# ========================================================================
#  工具函数：窗口居中
# ========================================================================


def _CenterOnScreen(Window: tk.Toplevel, Parent: Optional[tk.Widget] = None) -> None:
    """将窗口居中在父窗口（或屏幕）中央。
    如果父窗口隐藏（withdrawn），则直接居中于屏幕。"""
    Window.update_idletasks()
    Width = Window.winfo_width()
    Height = Window.winfo_height()
    if Parent is not None and Parent.winfo_viewable():
        ParentX = Parent.winfo_rootx()
        ParentY = Parent.winfo_rooty()
        ParentW = Parent.winfo_width()
        ParentH = Parent.winfo_height()
        X = ParentX + (ParentW - Width) // 2
        Y = ParentY + (ParentH - Height) // 2
    else:
        ScreenW = Window.winfo_screenwidth()
        ScreenH = Window.winfo_screenheight()
        X = (ScreenW - Width) // 2
        Y = (ScreenH - Height) // 2
    Window.geometry(f"+{X}+{Y}")


# ========================================================================
#  登录对话框
# ========================================================================

class LoginDialog(tk.Toplevel):
    """模态登录对话框。"""

    def __init__(self, Parent: tk.Tk, Db: MySqlManager) -> None:
        super().__init__(Parent)
        self._Db = Db
        self.Result: Optional[LoginUser] = None

        self.title("实验室考勤管理系统 - 登录")
        self.resizable(False, False)
        self.grab_set()

        Frame = ttk.Frame(self, padding=20)
        Frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(Frame, text="实验室考勤管理系统", font=("微软雅黑", 16, "bold")).pack(pady=(0, 20))

        ttk.Label(Frame, text="用户名：").pack(anchor=tk.W)
        self._UsernameVar = tk.StringVar()
        ttk.Entry(Frame, textvariable=self._UsernameVar, width=30).pack(fill=tk.X, pady=(0, 10))

        ttk.Label(Frame, text="密  码：").pack(anchor=tk.W)
        self._PasswordVar = tk.StringVar()
        ttk.Entry(Frame, textvariable=self._PasswordVar, width=30, show="*").pack(fill=tk.X, pady=(0, 20))

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.pack(fill=tk.X)
        ttk.Button(BtnFrame, text="登录", command=self._DoLogin).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(BtnFrame, text="退出", command=self._OnExitApp).pack(side=tk.RIGHT)

        self.bind("<Return>", lambda e: self._DoLogin())
        # 用户名和密码留空，由用户手动输入
        self._UsernameEntry = Frame.winfo_children()[2]
        self._UsernameEntry.focus_set()

        self.protocol("WM_DELETE_WINDOW", self._OnExitApp)
        _CenterOnScreen(self, self.master)
        self.wait_window()

    def _DoLogin(self) -> None:
        Username = self._UsernameVar.get().strip()
        Password = self._PasswordVar.get().strip()
        if not Username or not Password:
            messagebox.showwarning("提示", "请输入用户名和密码", parent=self)
            return
        try:
            UserData = self._Db.AuthenticateUser(Username, Password)
        except mysql.connector.Error as e:
            messagebox.showerror("数据库错误", f"无法连接数据库：{e}", parent=self)
            return
        if UserData is None:
            messagebox.showerror("登录失败", "用户名或密码错误，或账户已禁用", parent=self)
            return
        self.Result = LoginUser(
            UserId=UserData["user_id"],
            Username=UserData["username"],
            Role=UserData["role"],
            MemberId=UserData.get("member_id"),
            DisplayName=UserData.get("display_name", UserData["username"]),
        )
        self.destroy()

    def _OnExitApp(self) -> None:
        """退出登录并关闭整个程序。"""
        self.Result = None
        self.master.destroy()


# ========================================================================
#  UI 基础页面
# ========================================================================

class BasePage(ttk.Frame):
    """所有 Tab 页面的基类，提供统一的刷新接口。"""

    def __init__(self, Parent: ttk.Notebook, Db: MySqlManager, User: LoginUser) -> None:
        super().__init__(Parent)
        self._Db = Db
        self._User = User

    def Refresh(self) -> None:
        """子类重写此方法实现数据刷新。"""
        raise NotImplementedError


def _BuildTreeview(Parent: tk.Widget, Columns: List[str], Headings: List[str],
                   Widths: Optional[List[int]] = None) -> Tuple[ttk.Treeview, ttk.Scrollbar]:
    """创建带滚动条的 Treeview 控件。"""
    Frame = ttk.Frame(Parent)
    Frame.pack(fill=tk.BOTH, expand=True)

    Tree = ttk.Treeview(Frame, columns=Columns, show="headings", selectmode="browse")
    VScroll = ttk.Scrollbar(Frame, orient=tk.VERTICAL, command=Tree.yview)
    Tree.configure(yscrollcommand=VScroll.set)

    Tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    VScroll.pack(side=tk.RIGHT, fill=tk.Y)

    for i, (Col, Heading) in enumerate(zip(Columns, Headings)):
        Tree.heading(Col, text=Heading)
        Tree.column(Col, width=Widths[i] if Widths else 120, minwidth=80, anchor=tk.CENTER)

    return Tree, VScroll


def _FillTreeview(Tree: ttk.Treeview, Rows: List[Dict[str, Any]],
                  Columns: List[str]) -> None:
    """用查询结果填充 Treeview。"""
    for Item in Tree.get_children():
        Tree.delete(Item)
    for Row in Rows:
        Values = [str(Row.get(Col, "")) for Col in Columns]
        Tree.insert("", tk.END, values=Values)


def _GetSelectedRow(Tree: ttk.Treeview) -> Optional[Dict[str, str]]:
    """获取 Treeview 当前选中行，返回列名→值的字典。"""
    Selection = Tree.selection()
    if not Selection:
        return None
    Values = Tree.item(Selection[0], "values")
    Columns = Tree["columns"]
    return {Col: Values[i] if i < len(Values) else "" for i, Col in enumerate(Columns)}


def _Toolbar(Parent: tk.Widget) -> ttk.Frame:
    """创建一个工具栏 Frame。"""
    Bar = ttk.Frame(Parent)
    Bar.pack(fill=tk.X, pady=(0, 5))
    return Bar


def _AddRightClickMenu(Tree: ttk.Treeview, MenuItems: List[Tuple[str, Callable]]) -> None:
    """给 Treeview 添加右键菜单，MenuItems 为 (标签, 回调) 元组列表。"""
    Menu = tk.Menu(Tree, tearoff=0)
    for Label, Command in MenuItems:
        Menu.add_command(label=Label, command=Command)

    def OnRightClick(Event: tk.Event) -> None:
        Item = Tree.identify_row(Event.y)
        if Item:
            Tree.selection_set(Item)
            Menu.post(Event.x_root, Event.y_root)

    Tree.bind("<Button-3>", OnRightClick)


# ========================================================================
#  部门管理页
# ========================================================================

class DepartmentPage(BasePage):
    """部门信息管理。"""

    COLUMNS = ["department_id", "department_name", "description"]
    HEADINGS = ["编号", "部门名称", "描述"]
    WIDTHS = [60, 200, 300]

    def __init__(self, Parent: ttk.Notebook, Db: MySqlManager, User: LoginUser) -> None:
        super().__init__(Parent, Db, User)
        Bar = _Toolbar(self)
        ttk.Button(Bar, text="新增", command=self._OnAdd).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="修改", command=self._OnEdit).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="删除", command=self._OnDelete).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="刷新", command=self.Refresh).pack(side=tk.LEFT, padx=2)

        self._Tree, _ = _BuildTreeview(self, self.COLUMNS, self.HEADINGS, self.WIDTHS)
        _AddRightClickMenu(self._Tree, [("修改", self._OnEdit)])
        self.Refresh()

    def Refresh(self) -> None:
        try:
            Rows = self._Db.GetDepartments()
        except mysql.connector.Error:
            Rows = []
        _FillTreeview(self._Tree, Rows, self.COLUMNS)

    def _GetFormValues(self, Title: str, DeptId: Optional[int] = None,
                       Name: str = "", Desc: str = "") -> Optional[Tuple[Optional[int], str, str]]:
        """弹出编辑对话框，返回 (dept_id, name, description) 或 None。"""
        Dialog = tk.Toplevel(self)
        Dialog.title(Title)
        Dialog.resizable(False, False)
        Dialog.grab_set()

        Frame = ttk.Frame(Dialog, padding=15)
        Frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(Frame, text="部门名称：").grid(row=0, column=0, sticky=tk.W, pady=5)
        NameVar = tk.StringVar(value=Name)
        ttk.Entry(Frame, textvariable=NameVar, width=30).grid(row=0, column=1, pady=5)

        ttk.Label(Frame, text="描  述：").grid(row=1, column=0, sticky=tk.W, pady=5)
        DescVar = tk.StringVar(value=Desc)
        ttk.Entry(Frame, textvariable=DescVar, width=30).grid(row=1, column=1, pady=5)

        Result: List[Optional[Tuple[Optional[int], str, str]]] = [None]

        def OnOk() -> None:
            if not NameVar.get().strip():
                messagebox.showwarning("提示", "部门名称不能为空", parent=Dialog)
                return
            Result[0] = (DeptId, NameVar.get().strip(), DescVar.get().strip())
            Dialog.destroy()

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.grid(row=2, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(BtnFrame, text="确定", command=OnOk).pack(side=tk.RIGHT, padx=2)
        ttk.Button(BtnFrame, text="取消", command=Dialog.destroy).pack(side=tk.RIGHT)

        _CenterOnScreen(Dialog, self)
        Dialog.wait_window()
        return Result[0]

    def _OnAdd(self) -> None:
        Values = self._GetFormValues("新增部门")
        if Values is None:
            return
        _, Name, Desc = Values
        try:
            self._Db.UpsertDepartment(None, Name, Desc)
            self.Refresh()
        except mysql.connector.Error as e:
            messagebox.showerror("错误", f"新增失败：{e}")

    def _OnEdit(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        DeptId = int(Row.get("department_id", 0))
        Values = self._GetFormValues("修改部门", DeptId, Row.get("department_name", ""), Row.get("description", ""))
        if Values is None:
            return
        _, Name, Desc = Values
        try:
            self._Db.UpsertDepartment(DeptId, Name, Desc)
            self.Refresh()
        except mysql.connector.Error as e:
            messagebox.showerror("错误", f"修改失败：{e}")

    def _OnDelete(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        if not messagebox.askyesno("确认", f"确定删除部门「{Row.get('department_name', '')}」？"):
            return
        try:
            self._Db.DeleteDepartment(int(Row["department_id"]))
            self.Refresh()
            messagebox.showinfo("成功", "删除成功")
        except mysql.connector.Error as e:
            # 检查是否是外键约束错误
            if "1451" in str(e):
                messagebox.showerror("错误", "该部门下存在成员，无法删除。请先移除或重新分配成员所属部门。")
            else:
                messagebox.showerror("错误", f"删除失败：{e}")


# ========================================================================
#  成员管理页
# ========================================================================

class MemberPage(BasePage):
    """成员档案管理。"""

    COLUMNS = ["member_id", "name", "gender", "age", "identity", "position",
               "join_method", "contact", "department_name", "note"]
    HEADINGS = ["编号", "姓名", "性别", "年龄", "身份", "职位",
                "加入方式", "联系方式", "部门", "备注"]
    WIDTHS = [80, 80, 50, 50, 80, 100, 80, 120, 120, 150]

    def __init__(self, Parent: ttk.Notebook, Db: MySqlManager, User: LoginUser) -> None:
        super().__init__(Parent, Db, User)

        # 搜索栏
        SearchFrame = ttk.Frame(self)
        SearchFrame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(SearchFrame, text="编号/姓名：").pack(side=tk.LEFT)
        self._KeywordVar = tk.StringVar()
        ttk.Entry(SearchFrame, textvariable=self._KeywordVar, width=20).pack(side=tk.LEFT, padx=2)
        ttk.Button(SearchFrame, text="搜索", command=self.Refresh).pack(side=tk.LEFT, padx=2)
        ttk.Button(SearchFrame, text="清空", command=lambda: [self._KeywordVar.set(""), self.Refresh()]).pack(side=tk.LEFT)

        # 工具栏
        Bar = _Toolbar(self)
        ttk.Button(Bar, text="新增", command=self._OnAdd).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="修改", command=self._OnEdit).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="删除", command=self._OnDelete).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="导入 Excel", command=self._OnImport).pack(side=tk.LEFT, padx=2)

        self._Tree, _ = _BuildTreeview(self, self.COLUMNS, self.HEADINGS, self.WIDTHS)
        _AddRightClickMenu(self._Tree, [
            ("修改", self._OnEdit),
            ("删除", self._OnDelete),
        ])
        self.Refresh()

    def Refresh(self) -> None:
        Keyword = self._KeywordVar.get().strip()
        try:
            Rows = self._Db.GetMembers(Keyword=Keyword)
        except mysql.connector.Error:
            Rows = []
        _FillTreeview(self._Tree, Rows, self.COLUMNS)

    def _OnAdd(self) -> None:
        self._OpenForm("新增成员")

    def _OnEdit(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        self._OpenForm("修改成员", Row)

    def _OpenForm(self, Title: str, Row: Optional[Dict[str, str]] = None) -> None:
        """打开成员编辑表单。"""
        Dialog = tk.Toplevel(self)
        Dialog.title(Title)
        Dialog.resizable(False, False)
        Dialog.grab_set()

        Fields = ["member_id", "name", "gender", "age", "identity", "position",
                   "join_method", "contact", "note"]
        Labels = ["编号 *", "姓名 *", "性别", "年龄 *", "身份", "职位",
                   "加入方式", "联系方式", "备注"]
        Vars: Dict[str, tk.StringVar] = {}
        Frame = ttk.Frame(Dialog, padding=15)
        Frame.pack(fill=tk.BOTH, expand=True)

        for i, (Field, Label) in enumerate(zip(Fields, Labels)):
            ttk.Label(Frame, text=Label).grid(row=i, column=0, sticky=tk.W, pady=3)
            Vars[Field] = tk.StringVar(value=Row.get(Field, "") if Row else "")
            if Field == "gender":
                Combo = ttk.Combobox(Frame, textvariable=Vars[Field], values=GENDER_OPTIONS,
                                     state="readonly", width=27)
                Combo.grid(row=i, column=1, pady=3)
                if not Row:
                    Vars[Field].set(GENDER_OPTIONS[0])
            else:
                ttk.Entry(Frame, textvariable=Vars[Field], width=30).grid(row=i, column=1, pady=3)

        Result: List[Optional[Dict[str, str]]] = [None]

        def OnOk() -> None:
            Data = {F: V.get().strip() for F, V in Vars.items()}
            if not Data["member_id"] or not Data["name"] or not Data["age"]:
                messagebox.showwarning("提示", "编号、姓名、年龄为必填项", parent=Dialog)
                return
            if not IsValidAge(Data["age"]):
                messagebox.showwarning("提示", "年龄格式不正确（10-80 的数字）", parent=Dialog)
                return
            if Data["contact"] and not IsValidPhone(Data["contact"]):
                if not messagebox.askyesno("确认", "联系方式格式不正确，是否继续？", parent=Dialog):
                    return
            Result[0] = Data
            Dialog.destroy()

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.grid(row=len(Fields), column=0, columnspan=2, pady=(10, 0))
        ttk.Button(BtnFrame, text="确定", command=OnOk).pack(side=tk.RIGHT, padx=2)
        ttk.Button(BtnFrame, text="取消", command=Dialog.destroy).pack(side=tk.RIGHT)

        _CenterOnScreen(Dialog, self)
        Dialog.wait_window()
        Data = Result[0]
        if Data is None:
            return

        # 执行保存
        IsNew = Row is None
        MemberId = "" if IsNew else Row.get("member_id", "")
        try:
            if IsNew:
                self._Db.ExecuteNonQuery(
                    """
                    INSERT INTO members (member_id, name, gender, age, identity, position,
                        join_method, contact, note)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    tuple(Data[F] for F in Fields),
                )
                messagebox.showinfo("成功", "新增成员成功")
            else:
                self._Db.ExecuteNonQuery(
                    """
                    UPDATE members SET name=%s, gender=%s, age=%s, identity=%s, position=%s,
                        join_method=%s, contact=%s, note=%s
                    WHERE member_id = %s
                    """,
                    (Data["name"], Data["gender"], Data["age"], Data["identity"],
                     Data["position"], Data["join_method"], Data["contact"],
                     Data["note"], MemberId),
                )
                messagebox.showinfo("成功", "修改成员成功")
            self.Refresh()
        except mysql.connector.Error as e:
            # 检查是否是重复主键错误
            if "1062" in str(e):
                messagebox.showerror("错误", f"成员编号「{Data['member_id']}」已存在，请使用其他编号。")
            else:
                messagebox.showerror("错误", f"保存失败：{e}")

    def _OnDelete(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        if not messagebox.askyesno("确认", f"确定删除成员「{Row.get('name', '')}」？"):
            return
        try:
            self._Db.ExecuteNonQuery("DELETE FROM members WHERE member_id = %s", (Row["member_id"],))
            self.Refresh()
            messagebox.showinfo("成功", "删除成功")
        except mysql.connector.Error as e:
            # 检查是否是外键约束错误
            if "1451" in str(e):
                messagebox.showerror("错误", "该成员存在考勤记录，无法直接删除。请先删除相关考勤记录。")
            else:
                messagebox.showerror("错误", f"删除失败：{e}")

    def _OnImport(self) -> None:
        """从 Excel 文件批量导入成员（已废弃，保留用于兼容）。"""
        self._OnImportExcel()

    def _OnImportExcel(self) -> None:
        """从 Excel 文件批量导入成员。"""
        FilePath = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not FilePath:
            return
        
        try:
            Wb = load_workbook(FilePath, read_only=True)
            Ws = Wb.active
            Imported = 0
            Errors: List[str] = []
            
            for RowNum, Row in enumerate(Ws.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if Row[0] is None:
                    continue
                
                # 提取并验证数据
                MemberId = str(Row[0]).strip() if Row[0] is not None else ""
                Name = str(Row[1]).strip() if Row[1] is not None else ""
                Gender = str(Row[2]).strip() if Row[2] is not None else "男"
                AgeText = str(Row[3]).strip() if Row[3] is not None else ""
                Identity = str(Row[4]).strip() if Row[4] is not None else ""
                Position = str(Row[5]).strip() if Row[5] is not None else ""
                JoinMethod = str(Row[6]).strip() if Row[6] is not None else ""
                Contact = str(Row[7]).strip() if Row[7] is not None else ""
                Note = str(Row[8]).strip() if Row[8] is not None else ""
                
                # 必填字段检查
                if not MemberId or not Name:
                    Errors.append(f"第{RowNum}行：成员编号和姓名为必填项")
                    continue
                
                # 手机号格式校验
                if Contact and not IsValidPhone(Contact):
                    Errors.append(f"第{RowNum}行：手机号格式不正确 ({Contact})")
                    continue
                
                # 年龄范围校验
                if AgeText and not IsValidAge(AgeText):
                    Errors.append(f"第{RowNum}行：年龄应在10-80之间 ({AgeText})")
                    continue
                
                # 性别校验
                if Gender not in GENDER_OPTIONS:
                    Errors.append(f"第{RowNum}行：性别应为'男'或'女' ({Gender})")
                    continue
                
                try:
                    self._Db.ExecuteNonQuery(
                        """
                        INSERT INTO members (member_id, name, gender, age, identity, position,
                            join_method, contact, note)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (MemberId, Name, Gender, int(AgeText) if AgeText else 0, Identity,
                         Position, JoinMethod, Contact, Note),
                    )
                    Imported += 1
                except mysql.connector.Error as e:
                    Errors.append(f"第{RowNum}行：数据库错误 - {e}")
            
            Wb.close()
            
            # 显示结果
            if Errors:
                ErrorMsg = f"成功导入 {Imported} 条记录\n\n以下 {len(Errors)} 条记录导入失败：\n" + "\n".join(Errors[:10])
                if len(Errors) > 10:
                    ErrorMsg += f"\n...还有 {len(Errors) - 10} 条错误未显示"
                messagebox.showwarning("导入完成（部分失败）", ErrorMsg)
            else:
                messagebox.showinfo("导入完成", f"成功导入 {Imported} 条记录")
            
            self.Refresh()
        except Exception as e:
            messagebox.showerror("导入失败", f"读取文件时出错：{e}")

    def _OnImportCsv(self) -> None:
        """从 CSV 文件批量导入成员。"""
        FilePath = filedialog.askopenfilename(
            title="选择 CSV 文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not FilePath:
            return
        
        try:
            with open(FilePath, 'r', encoding='utf-8-sig') as F:
                Reader = csv.DictReader(F)
                Imported = 0
                Errors: List[str] = []
                
                for RowNum, Row in enumerate(Reader, start=2):
                    # 提取并验证数据
                    MemberId = NormalizeText(Row.get('member_id', ''))
                    Name = NormalizeText(Row.get('name', ''))
                    Gender = NormalizeText(Row.get('gender', '男'))
                    AgeText = NormalizeText(Row.get('age', ''))
                    Identity = NormalizeText(Row.get('identity', ''))
                    Position = NormalizeText(Row.get('position', ''))
                    JoinMethod = NormalizeText(Row.get('join_method', ''))
                    Contact = NormalizeText(Row.get('contact', ''))
                    Note = NormalizeText(Row.get('note', ''))
                    
                    # 必填字段检查
                    if not MemberId or not Name:
                        Errors.append(f"第{RowNum}行：成员编号和姓名为必填项")
                        continue
                    
                    # 手机号格式校验
                    if Contact and not IsValidPhone(Contact):
                        Errors.append(f"第{RowNum}行：手机号格式不正确 ({Contact})")
                        continue
                    
                    # 年龄范围校验
                    if AgeText and not IsValidAge(AgeText):
                        Errors.append(f"第{RowNum}行：年龄应在10-80之间 ({AgeText})")
                        continue
                    
                    # 性别校验
                    if Gender not in GENDER_OPTIONS:
                        Errors.append(f"第{RowNum}行：性别应为'男'或'女' ({Gender})")
                        continue
                    
                    try:
                        self._Db.ExecuteNonQuery(
                            """
                            INSERT INTO members (member_id, name, gender, age, identity, position,
                                join_method, contact, note)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (MemberId, Name, Gender, int(AgeText) if AgeText else 0, Identity,
                             Position, JoinMethod, Contact, Note),
                        )
                        Imported += 1
                    except mysql.connector.Error as e:
                        Errors.append(f"第{RowNum}行：数据库错误 - {e}")
            
            # 显示结果
            if Errors:
                ErrorMsg = f"成功导入 {Imported} 条记录\n\n以下 {len(Errors)} 条记录导入失败：\n" + "\n".join(Errors[:10])
                if len(Errors) > 10:
                    ErrorMsg += f"\n...还有 {len(Errors) - 10} 条错误未显示"
                messagebox.showwarning("导入完成（部分失败）", ErrorMsg)
            else:
                messagebox.showinfo("导入完成", f"成功导入 {Imported} 条记录")
            
            self.Refresh()
        except Exception as e:
            messagebox.showerror("导入失败", f"读取文件时出错：{e}")


# ========================================================================
#  考勤管理页
# ========================================================================

class AttendancePage(BasePage):
    """考勤记录管理。"""

    COLUMNS = ["attendance_id", "member_id", "member_name", "department_name",
               "attendance_date", "status", "check_in_time", "check_out_time", "remark"]
    HEADINGS = ["编号", "成员编号", "成员姓名", "部门",
                "日期", "状态", "签到时间", "签退时间", "备注"]
    WIDTHS = [60, 80, 80, 100, 100, 70, 90, 90, 150]

    def __init__(self, Parent: ttk.Notebook, Db: MySqlManager, User: LoginUser) -> None:
        super().__init__(Parent, Db, User)

        # 过滤栏
        FilterFrame = ttk.Frame(self)
        FilterFrame.pack(fill=tk.X, pady=(0, 5))

        # 学生角色：自动限制只能查看自己的考勤记录
        if self._User.Role == ROLE_STUDENT and self._User.MemberId:
            self._MemberVar = tk.StringVar(value=self._User.MemberId)
            MemberEntry = ttk.Entry(FilterFrame, textvariable=self._MemberVar, width=12, state="readonly")
            MemberEntry.pack(side=tk.LEFT, padx=2)
            ttk.Label(FilterFrame, text="（仅查看本人）", foreground="gray").pack(side=tk.LEFT, padx=2)
        else:
            ttk.Label(FilterFrame, text="成员编号：").pack(side=tk.LEFT)
            self._MemberVar = tk.StringVar()
            ttk.Entry(FilterFrame, textvariable=self._MemberVar, width=12).pack(side=tk.LEFT, padx=2)

        # 开始日期：年/月/日下拉栏
        ttk.Label(FilterFrame, text="开始：").pack(side=tk.LEFT)
        self._StartYearVar = tk.StringVar()
        self._StartMonthVar = tk.StringVar()
        self._StartDayVar = tk.StringVar()
        StartFrame = ttk.Frame(FilterFrame)
        StartFrame.pack(side=tk.LEFT, padx=2)
        YearValues = [""] + [str(y) for y in range(2020, 2036)]
        MonthValues = [""] + [f"{m:02d}" for m in range(1, 13)]
        DayValues = [""] + [f"{d:02d}" for d in range(1, 32)]
        ttk.Combobox(StartFrame, textvariable=self._StartYearVar, values=YearValues,
                     width=6, state="readonly").pack(side=tk.LEFT)
        ttk.Label(StartFrame, text="-").pack(side=tk.LEFT, padx=1)
        ttk.Combobox(StartFrame, textvariable=self._StartMonthVar, values=MonthValues,
                     width=4, state="readonly").pack(side=tk.LEFT)
        ttk.Label(StartFrame, text="-").pack(side=tk.LEFT, padx=1)
        ttk.Combobox(StartFrame, textvariable=self._StartDayVar, values=DayValues,
                     width=4, state="readonly").pack(side=tk.LEFT)

        # 结束日期：年/月/日下拉栏
        ttk.Label(FilterFrame, text="结束：").pack(side=tk.LEFT)
        self._EndYearVar = tk.StringVar()
        self._EndMonthVar = tk.StringVar()
        self._EndDayVar = tk.StringVar()
        EndFrame = ttk.Frame(FilterFrame)
        EndFrame.pack(side=tk.LEFT, padx=2)
        ttk.Combobox(EndFrame, textvariable=self._EndYearVar, values=YearValues,
                     width=6, state="readonly").pack(side=tk.LEFT)
        ttk.Label(EndFrame, text="-").pack(side=tk.LEFT, padx=1)
        ttk.Combobox(EndFrame, textvariable=self._EndMonthVar, values=MonthValues,
                     width=4, state="readonly").pack(side=tk.LEFT)
        ttk.Label(EndFrame, text="-").pack(side=tk.LEFT, padx=1)
        ttk.Combobox(EndFrame, textvariable=self._EndDayVar, values=DayValues,
                     width=4, state="readonly").pack(side=tk.LEFT)

        ttk.Label(FilterFrame, text="状态：").pack(side=tk.LEFT)
        self._StatusVar = tk.StringVar()
        StatusCombo = ttk.Combobox(FilterFrame, textvariable=self._StatusVar,
                                   values=[""] + ATTENDANCE_STATUSES, width=8)
        StatusCombo.pack(side=tk.LEFT, padx=2)
        ttk.Button(FilterFrame, text="查询", command=self.Refresh).pack(side=tk.LEFT, padx=2)
        ttk.Button(FilterFrame, text="清空", command=self._ClearFilter).pack(side=tk.LEFT, padx=2)

        Bar = _Toolbar(self)
        # 仅管理员可以增删改考勤记录
        if self._User.Role == ROLE_ADMIN:
            ttk.Button(Bar, text="新增", command=self._OnAdd).pack(side=tk.LEFT, padx=2)
            ttk.Button(Bar, text="修改", command=self._OnEdit).pack(side=tk.LEFT, padx=2)
            ttk.Button(Bar, text="删除", command=self._OnDelete).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="导出 CSV", command=self._OnExportCsv).pack(side=tk.LEFT, padx=2)

        self._Tree, _ = _BuildTreeview(self, self.COLUMNS, self.HEADINGS, self.WIDTHS)
        MenuItems = []
        if self._User.Role == ROLE_ADMIN:
            MenuItems = [("修改", self._OnEdit), ("删除", self._OnDelete)]
        _AddRightClickMenu(self._Tree, MenuItems)
        self.Refresh()

    def Refresh(self) -> None:
        # 组装开始/结束日期
        StartDate = self._BuildDateStr(self._StartYearVar, self._StartMonthVar, self._StartDayVar)
        EndDate = self._BuildDateStr(self._EndYearVar, self._EndMonthVar, self._EndDayVar)

        # 校验：结束日期不得早于开始日期
        if StartDate and EndDate and EndDate < StartDate:
            messagebox.showwarning("提示", "结束日期不得早于开始日期")
            return

        try:
            Rows = self._Db.GetAttendanceRecords(
                MemberId=self._MemberVar.get().strip() or None,
                StartDate=StartDate,
                EndDate=EndDate,
                Status=self._StatusVar.get().strip(),
            )
        except mysql.connector.Error as e:
            # 检查是否是数据库连接错误
            if "2003" in str(e) or "2002" in str(e):
                messagebox.showerror("数据库错误", "无法连接到数据库，请检查 MySQL 服务是否正在运行。")
            else:
                messagebox.showerror("查询失败", f"查询数据时出错：{e}")
            Rows = []
        _FillTreeview(self._Tree, Rows, self.COLUMNS)

    @staticmethod
    def _BuildDateStr(YearVar: tk.StringVar, MonthVar: tk.StringVar, DayVar: tk.StringVar) -> str:
        """从年/月/日 StringVar 组装日期字符串。"""
        Year = YearVar.get()
        Month = MonthVar.get()
        Day = DayVar.get()
        if Year and Month and Day:
            return f"{Year}-{Month}-{Day}"
        if Year and Month:
            return f"{Year}-{Month}"
        if Year:
            return Year
        return ""

    def _ClearFilter(self) -> None:
        """清空所有过滤条件并刷新。"""
        self._MemberVar.set("")
        self._StartYearVar.set("")
        self._StartMonthVar.set("")
        self._StartDayVar.set("")
        self._EndYearVar.set("")
        self._EndMonthVar.set("")
        self._EndDayVar.set("")
        self._StatusVar.set("")
        self.Refresh()

    def _OpenForm(self, Title: str, Row: Optional[Dict[str, str]] = None) -> None:
        Dialog = tk.Toplevel(self)
        Dialog.title(Title)
        Dialog.resizable(False, False)
        Dialog.grab_set()

        Frame = ttk.Frame(Dialog, padding=15)
        Frame.pack(fill=tk.BOTH, expand=True)

        RowIdx = 0

        # 成员编号
        ttk.Label(Frame, text="成员编号 *：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        MemberIdVar = tk.StringVar(value=Row.get("member_id", "") if Row else "")
        ttk.Entry(Frame, textvariable=MemberIdVar, width=30).grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        # 日期：年 / 月 / 日 下拉栏
        ttk.Label(Frame, text="日期 *：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        DateFrame = ttk.Frame(Frame)
        DateFrame.grid(row=RowIdx, column=1, sticky=tk.W, pady=3)

        Now = datetime.now()
        CurrentYear = Now.year
        # 解析已有值
        DefaultDate = Row.get("attendance_date", "") if Row else ""
        if DefaultDate:
            Parts = DefaultDate.split("-")
            DefaultYear = Parts[0] if len(Parts) > 0 else str(CurrentYear)
            DefaultMonth = Parts[1] if len(Parts) > 1 else "01"
            DefaultDay = Parts[2] if len(Parts) > 2 else "01"
        else:
            DefaultYear = str(CurrentYear)
            DefaultMonth = f"{Now.month:02d}"
            DefaultDay = f"{Now.day:02d}"

        YearVar = tk.StringVar(value=DefaultYear)
        YearCombo = ttk.Combobox(DateFrame, textvariable=YearVar, values=[str(y) for y in range(2020, 2036)],
                                  width=6, state="readonly")
        YearCombo.pack(side=tk.LEFT)
        ttk.Label(DateFrame, text="年").pack(side=tk.LEFT, padx=2)

        MonthVar = tk.StringVar(value=DefaultMonth)
        MonthCombo = ttk.Combobox(DateFrame, textvariable=MonthVar, values=[f"{m:02d}" for m in range(1, 13)],
                                   width=4, state="readonly")
        MonthCombo.pack(side=tk.LEFT)
        ttk.Label(DateFrame, text="月").pack(side=tk.LEFT, padx=2)

        DayVar = tk.StringVar(value=DefaultDay)
        DayCombo = ttk.Combobox(DateFrame, textvariable=DayVar, values=[f"{d:02d}" for d in range(1, 32)],
                                 width=4, state="readonly")
        DayCombo.pack(side=tk.LEFT)
        ttk.Label(DateFrame, text="日").pack(side=tk.LEFT, padx=2)
        RowIdx += 1

        # 状态
        ttk.Label(Frame, text="状态 *：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        StatusVar = tk.StringVar(value=Row.get("status", "") if Row else ATTENDANCE_STATUSES[0])
        StatusCombo = ttk.Combobox(Frame, textvariable=StatusVar, values=ATTENDANCE_STATUSES,
                                    state="readonly", width=27)
        StatusCombo.grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        # 签到时间：时 / 分 下拉栏
        ttk.Label(Frame, text="签到时间：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        InFrame = ttk.Frame(Frame)
        InFrame.grid(row=RowIdx, column=1, sticky=tk.W, pady=3)

        DefaultIn = Row.get("check_in_time", "") if Row else ""
        if DefaultIn:
            InParts = DefaultIn.split(":")
            DefaultInHour = InParts[0]
            DefaultInMin = InParts[1] if len(InParts) > 1 else "00"
        else:
            DefaultInHour = ""
            DefaultInMin = ""

        InHourVar = tk.StringVar(value=DefaultInHour)
        InHourCombo = ttk.Combobox(InFrame, textvariable=InHourVar,
                                    values=[""] + [f"{h:02d}" for h in range(0, 24)],
                                    width=4, state="readonly")
        InHourCombo.pack(side=tk.LEFT)
        ttk.Label(InFrame, text="时").pack(side=tk.LEFT, padx=2)

        InMinVar = tk.StringVar(value=DefaultInMin)
        InMinCombo = ttk.Combobox(InFrame, textvariable=InMinVar,
                                   values=[""] + [f"{m:02d}" for m in range(0, 60)],
                                   width=4, state="readonly")
        InMinCombo.pack(side=tk.LEFT)
        ttk.Label(InFrame, text="分").pack(side=tk.LEFT, padx=2)
        RowIdx += 1

        # 签退时间：时 / 分 下拉栏
        ttk.Label(Frame, text="签退时间：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        OutFrame = ttk.Frame(Frame)
        OutFrame.grid(row=RowIdx, column=1, sticky=tk.W, pady=3)

        DefaultOut = Row.get("check_out_time", "") if Row else ""
        if DefaultOut:
            OutParts = DefaultOut.split(":")
            DefaultOutHour = OutParts[0]
            DefaultOutMin = OutParts[1] if len(OutParts) > 1 else "00"
        else:
            DefaultOutHour = ""
            DefaultOutMin = ""

        OutHourVar = tk.StringVar(value=DefaultOutHour)
        OutHourCombo = ttk.Combobox(OutFrame, textvariable=OutHourVar,
                                     values=[""] + [f"{h:02d}" for h in range(0, 24)],
                                     width=4, state="readonly")
        OutHourCombo.pack(side=tk.LEFT)
        ttk.Label(OutFrame, text="时").pack(side=tk.LEFT, padx=2)

        OutMinVar = tk.StringVar(value=DefaultOutMin)
        OutMinCombo = ttk.Combobox(OutFrame, textvariable=OutMinVar,
                                    values=[""] + [f"{m:02d}" for m in range(0, 60)],
                                    width=4, state="readonly")
        OutMinCombo.pack(side=tk.LEFT)
        ttk.Label(OutFrame, text="分").pack(side=tk.LEFT, padx=2)
        RowIdx += 1

        # 备注
        ttk.Label(Frame, text="备注：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        RemarkVar = tk.StringVar(value=Row.get("remark", "") if Row else "")
        ttk.Entry(Frame, textvariable=RemarkVar, width=30).grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        Result: List[Optional[Dict[str, str]]] = [None]

        def OnOk() -> None:
            # 组装日期
            DateStr = f"{YearVar.get()}-{MonthVar.get()}-{DayVar.get()}"
            # 组装时间
            InHour = InHourVar.get()
            InMin = InMinVar.get()
            InTime = f"{InHour}:{InMin}:00" if InHour and InMin else ""
            OutHour = OutHourVar.get()
            OutMin = OutMinVar.get()
            OutTime = f"{OutHour}:{OutMin}:00" if OutHour and OutMin else ""

            # 校验：签退时间不能早于签到时间
            if InTime and OutTime and OutTime <= InTime:
                messagebox.showwarning("提示", "签退时间不能早于或等于签到时间", parent=Dialog)
                return

            Data = {
                "member_id": MemberIdVar.get().strip(),
                "attendance_date": DateStr,
                "status": StatusVar.get(),
                "check_in_time": InTime,
                "check_out_time": OutTime,
                "remark": RemarkVar.get().strip(),
            }
            if not Data["member_id"] or not Data["attendance_date"] or not Data["status"]:
                messagebox.showwarning("提示", "成员编号、日期、状态为必填项", parent=Dialog)
                return
            Result[0] = Data
            Dialog.destroy()

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.grid(row=RowIdx, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(BtnFrame, text="确定", command=OnOk).pack(side=tk.RIGHT, padx=2)
        ttk.Button(BtnFrame, text="取消", command=Dialog.destroy).pack(side=tk.RIGHT)

        Dialog.bind("<Return>", lambda e: OnOk())
        _CenterOnScreen(Dialog, self)
        Dialog.wait_window()
        Data = Result[0]
        if Data is None:
            return

        AttendanceId = None if Row is None else int(Row.get("attendance_id", 0))
        try:
            self._Db.UpsertAttendance(
                AttendanceId,
                Data["member_id"], Data["attendance_date"], Data["status"],
                Data["check_in_time"] or None,
                Data["check_out_time"] or None,
                Data["remark"],
            )
            messagebox.showinfo("成功", "保存成功")
            self.Refresh()
        except mysql.connector.Error as e:
            # 检查是否是重复唯一键错误
            if "1062" in str(e):
                messagebox.showerror("错误", f"成员「{Data['member_id']}」在日期「{Data['attendance_date']}」已有考勤记录，请先删除或修改原记录。")
            # 检查是否是外键约束错误（成员不存在）
            elif "1452" in str(e):
                messagebox.showerror("错误", f"成员编号「{Data['member_id']}」不存在，请先添加该成员。")
            else:
                messagebox.showerror("错误", f"保存失败：{e}")

    def _OnAdd(self) -> None:
        self._OpenForm("新增考勤")

    def _OnEdit(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        self._OpenForm("修改考勤", Row)

    def _OnDelete(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        if not messagebox.askyesno("确认", "确定删除该考勤记录？"):
            return
        try:
            self._Db.DeleteAttendance(int(Row["attendance_id"]))
            self.Refresh()
            messagebox.showinfo("成功", "删除成功")
        except mysql.connector.Error as e:
            messagebox.showerror("错误", f"删除失败：{e}")

    def _OnExportCsv(self) -> None:
        FilePath = filedialog.asksaveasfilename(
            title="导出 CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
        )
        if not FilePath:
            return
        try:
            Rows = self._Db.GetAttendanceRecords()
            with open(FilePath, "w", newline="", encoding="utf-8-sig") as F:
                Writer = csv.DictWriter(F, fieldnames=self.COLUMNS)
                Writer.writeheader()
                Writer.writerows(Rows)
            messagebox.showinfo("导出完成", f"已导出 {len(Rows)} 条记录到 {FilePath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))


# ========================================================================
#  用户管理页
# ========================================================================

class UserPage(BasePage):
    """用户账户管理。"""

    COLUMNS = ["user_id", "username", "role", "member_id", "member_name", "is_active"]
    HEADINGS = ["编号", "用户名", "角色", "关联成员编号", "成员姓名", "启用"]
    WIDTHS = [60, 120, 80, 100, 100, 50]

    def __init__(self, Parent: ttk.Notebook, Db: MySqlManager, User: LoginUser) -> None:
        super().__init__(Parent, Db, User)
        Bar = _Toolbar(self)
        ttk.Button(Bar, text="新增", command=self._OnAdd).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="修改", command=self._OnEdit).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="删除", command=self._OnDelete).pack(side=tk.LEFT, padx=2)
        if self._User.Role == ROLE_ADMIN:
            ttk.Button(Bar, text="重置密码", command=self._OnResetPassword).pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="刷新", command=self.Refresh).pack(side=tk.LEFT, padx=2)

        self._Tree, _ = _BuildTreeview(self, self.COLUMNS, self.HEADINGS, self.WIDTHS)
        MenuItems = [("修改", self._OnEdit), ("删除", self._OnDelete)]
        if self._User.Role == ROLE_ADMIN:
            MenuItems.append(("重置密码", self._OnResetPassword))
        _AddRightClickMenu(self._Tree, MenuItems)
        self.Refresh()

    def Refresh(self) -> None:
        try:
            Rows = self._Db.GetUsers()
        except mysql.connector.Error:
            Rows = []
        # 转换 is_active 为显示文字
        for Row in Rows:
            Row["is_active"] = "是" if Row.get("is_active") else "否"
        _FillTreeview(self._Tree, Rows, self.COLUMNS)

    def _OpenForm(self, Title: str, Row: Optional[Dict[str, str]] = None) -> None:
        Dialog = tk.Toplevel(self)
        Dialog.title(Title)
        Dialog.resizable(False, False)
        Dialog.grab_set()

        Frame = ttk.Frame(Dialog, padding=15)
        Frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(Frame, text="用户名 *：").grid(row=0, column=0, sticky=tk.W, pady=3)
        UsernameVar = tk.StringVar(value=Row.get("username", "") if Row else "")
        ttk.Entry(Frame, textvariable=UsernameVar, width=30).grid(row=0, column=1, pady=3)

        ttk.Label(Frame, text="密码 *：").grid(row=1, column=0, sticky=tk.W, pady=3)
        PasswordVar = tk.StringVar()
        PwdEntry = ttk.Entry(Frame, textvariable=PasswordVar, width=30, show="*")
        PwdEntry.grid(row=1, column=1, pady=3)
        if Row:
            PwdEntry.configure(state="readonly")

        ttk.Label(Frame, text="角色：").grid(row=2, column=0, sticky=tk.W, pady=3)
        RoleVar = tk.StringVar(value=Row.get("role", ROLE_STUDENT) if Row else ROLE_STUDENT)
        RoleValues = [ROLE_ADMIN, ROLE_STAFF, ROLE_STUDENT]
        # 普通老师只能创建/修改学生账号
        if self._User.Role == ROLE_STAFF:
            RoleValues = [ROLE_STUDENT]
        RoleCombo = ttk.Combobox(Frame, textvariable=RoleVar,
                                  values=RoleValues,
                                  state="readonly", width=27)
        RoleCombo.grid(row=2, column=1, pady=3)

        ttk.Label(Frame, text="关联成员：").grid(row=3, column=0, sticky=tk.W, pady=3)
        MemberIdVar = tk.StringVar(value=Row.get("member_id", "") if Row else "")
        ttk.Entry(Frame, textvariable=MemberIdVar, width=30).grid(row=3, column=1, pady=3)

        ttk.Label(Frame, text="启用：").grid(row=4, column=0, sticky=tk.W, pady=3)
        ActiveVar = tk.BooleanVar(value=Row.get("is_active", "是") == "是" if Row else True)
        ttk.Checkbutton(Frame, variable=ActiveVar).grid(row=4, column=1, sticky=tk.W, pady=3)

        Result: List[Optional[Dict[str, Any]]] = [None]

        def OnOk() -> None:
            if not UsernameVar.get().strip():
                messagebox.showwarning("提示", "用户名不能为空", parent=Dialog)
                return
            if Row is None and not PasswordVar.get().strip():
                messagebox.showwarning("提示", "密码不能为空", parent=Dialog)
                return
            Result[0] = {
                "username": UsernameVar.get().strip(),
                "password": PasswordVar.get().strip(),
                "role": RoleVar.get(),
                "member_id": MemberIdVar.get().strip() or None,
                "is_active": ActiveVar.get(),
            }
            Dialog.destroy()

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.grid(row=5, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(BtnFrame, text="确定", command=OnOk).pack(side=tk.RIGHT, padx=2)
        ttk.Button(BtnFrame, text="取消", command=Dialog.destroy).pack(side=tk.RIGHT)

        _CenterOnScreen(Dialog, self)
        Dialog.wait_window()
        Data = Result[0]
        if Data is None:
            return

        UserId = None if Row is None else int(Row.get("user_id", 0))
        try:
            if Row is None:
                self._Db.UpsertUser(
                    None, Data["username"], Data["password"],
                    Data["role"], Data["member_id"], Data["is_active"],
                )
            else:
                # 更新时不改密码
                self._Db.ExecuteNonQuery(
                    """
                    UPDATE user_accounts
                    SET username = %s, role = %s, member_id = %s, is_active = %s
                    WHERE user_id = %s
                    """,
                    (Data["username"], Data["role"], Data["member_id"],
                     int(Data["is_active"]), UserId),
                )
            self.Refresh()
        except mysql.connector.Error as e:
            messagebox.showerror("错误", f"保存失败：{e}")

    def _OnAdd(self) -> None:
        self._OpenForm("新增用户")

    def _OnEdit(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        # Staff 不能修改 Admin 或其他 Staff 账号
        if self._User.Role == ROLE_STAFF and Row.get("role") in (ROLE_ADMIN, ROLE_STAFF):
            messagebox.showwarning("权限不足", "您没有权限修改管理员或其他老师的账号")
            return
        self._OpenForm("修改用户", Row)

    def _OnDelete(self) -> None:
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return
        # Staff 不能删除 Admin 或其他 Staff 账号
        if self._User.Role == ROLE_STAFF and Row.get("role") in (ROLE_ADMIN, ROLE_STAFF):
            messagebox.showwarning("权限不足", "您没有权限删除管理员或其他老师的账号")
            return
        if not messagebox.askyesno("确认", f"确定删除用户「{Row.get('username', '')}」？"):
            return
        try:
            self._Db.DeleteUser(int(Row["user_id"]))
            self.Refresh()
        except mysql.connector.Error as e:
            messagebox.showerror("错误", f"删除失败：{e}")

    def _OnResetPassword(self) -> None:
        """管理员重置选中用户的密码为固定值 123456。"""
        Row = _GetSelectedRow(self._Tree)
        if Row is None:
            messagebox.showinfo("提示", "请先选择一条记录")
            return

        Username = Row.get("username", "")
        if not messagebox.askyesno("确认重置密码",
                                    f"确定将用户「{Username}」的密码重置为 123456？",
                                    parent=self):
            return

        try:
            self._Db.AdminResetPassword(int(Row["user_id"]), "123456")
            messagebox.showinfo("成功", f"用户「{Username}」的密码已重置为 123456")
        except Exception as e:
            messagebox.showerror("错误", f"重置失败：{e}")


# ========================================================================
#  统计图表页
# ========================================================================

class StatsPage(BasePage):
    """考勤统计图表。"""

    def __init__(self, Parent: ttk.Notebook, Db: MySqlManager, User: LoginUser) -> None:
        super().__init__(Parent, Db, User)

        Bar = _Toolbar(self)
        ttk.Label(Bar, text="统计方式：").pack(side=tk.LEFT)
        self._ModeVar = tk.StringVar(value="status")
        ModeCombo = ttk.Combobox(Bar, textvariable=self._ModeVar,
                                  values=["status", "department"], state="readonly", width=15)
        ModeCombo.pack(side=tk.LEFT, padx=2)
        ttk.Button(Bar, text="刷新图表", command=self.Refresh).pack(side=tk.LEFT, padx=5)

        # 图表容器
        self._Fig = Figure(figsize=(8, 4), dpi=100)
        self._Ax = self._Fig.add_subplot(111)
        self._Canvas = FigureCanvasTkAgg(self._Fig, master=self)
        self._Canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=10)

        self.Refresh()

    def Refresh(self) -> None:
        Mode = self._ModeVar.get()
        self._Ax.clear()
        try:
            if Mode == "status":
                Rows = self._Db.ExecuteQuery(
                    "SELECT status, COUNT(*) AS Cnt FROM attendance_records GROUP BY status ORDER BY Cnt DESC"
                )
                Labels = [R["status"] for R in Rows]
                Values = [R["Cnt"] for R in Rows]
                self._Ax.bar(Labels, Values, color=["#4CAF50", "#FF9800", "#F44336", "#2196F3"])
                self._Ax.set_title("考勤状态分布")
            else:
                Rows = self._Db.ExecuteQuery(
                    """
                    SELECT IFNULL(d.department_name, '未分配') AS dept, COUNT(a.attendance_id) AS Cnt
                    FROM attendance_records a
                    JOIN members m ON a.member_id = m.member_id
                    LEFT JOIN departments d ON m.department_id = d.department_id
                    GROUP BY d.department_id
                    ORDER BY Cnt DESC
                    """
                )
                Labels = [R["dept"] for R in Rows]
                Values = [R["Cnt"] for R in Rows]
                self._Ax.bar(Labels, Values, color=["#3F51B5", "#009688", "#FF5722", "#607D8B"])
                self._Ax.set_title("各部门考勤记录数")
            self._Ax.set_ylabel("数量")
            self._Ax.tick_params(axis="x", rotation=15)
            self._Fig.tight_layout()
            self._Canvas.draw()
        except mysql.connector.Error:
            self._Ax.text(0.5, 0.5, "暂无数据", ha="center", va="center", fontsize=14)
            self._Canvas.draw()


# ========================================================================
#  主应用
# ========================================================================

class Application(tk.Tk):
    """实验室考勤管理系统主应用。"""

    def __init__(self) -> None:
        super().__init__()
        self.title("实验室考勤管理系统")
        self.geometry("1100x700")
        self.minsize(900, 600)

        self._Db: Optional[MySqlManager] = None
        self._User: Optional[LoginUser] = None
        self._Notebook: Optional[ttk.Notebook] = None

        # 登录前隐藏主窗口
        self.withdraw()
        # 主窗口居中
        _CenterOnScreen(self)

    def destroy(self) -> None:
        """重写 destroy，确保进程结束时关闭数据库连接。"""
        if self._Db is not None:
            try:
                self._Db.Close()
            except Exception:
                pass
        super().destroy()

    def Run(self) -> None:
        """启动应用。"""
        # 尝试自动启动 MySQL 服务
        mysql_ok = self._EnsureMySqlRunning()
        print(f"[启动] mysql_ok={mysql_ok}")

        # 连接数据库，失败时弹出配置对话框让用户修改
        while True:
            try:
                if not mysql_ok:
                    # 自动启动失败，立即弹出配置窗口（跳过连接尝试）
                    print("[启动] MySQL 未就绪，准备弹出配置窗口...")
                    Config = self._DbConfigDialog(
                        "MySQL 服务无法自动启动（数据目录未初始化）。\n"
                        "请先手动初始化 MySQL 数据目录，\n"
                        "或修改以下数据库连接参数后重试。"
                    )
                else:
                    self._InitDatabase()
                    break  # 连接成功
            except Exception as e:
                print(f"[启动] 数据库初始失败: {e}")
                Config = self._DbConfigDialog(
                    f"无法连接数据库：{e}\n请检查以下配置是否正确。"
                )
            if Config is None:
                print("[启动] 用户取消配置，退出")
                self.destroy()  # 用户取消
                return
            # 用用户输入的配置重试
            print(f"[启动] 使用新配置重试: {Config.Host}:{Config.Port}")
            self._Db = MySqlManager(Config)
            try:
                self._Db.EnsureSchema()
                break  # 重试成功
            except Exception:
                print("[启动] 重试仍失败，继续弹窗")
                mysql_ok = False  # 再次失败，继续弹窗
                continue

        # 显示登录对话框
        LoginDlg = LoginDialog(self, self._Db)
        if LoginDlg.Result is None:
            self.destroy()
            return
        self._User = LoginDlg.Result
        # 登录成功后显示主窗口
        self.deiconify()
        self.title(f"实验室考勤管理系统 - {self._User.DisplayName} ({self._User.Role})")
        self._BuildMainWindow()
        # 启动数据库连接健康检查（每60秒）
        self._StartDbHealthCheck()

        try:
            self.mainloop()
        finally:
            if self._Db is not None:
                self._Db.Close()

    def _EnsureMySqlRunning(self) -> bool:
        """确保 MySQL 服务正在运行。如果未运行则尝试自动启动。
        返回 True 表示 MySQL 已就绪，False 表示无法启动。"""
        if start_mysql is None:
            # mysql_service 模块不可用，跳过自动启动
            return True

        try:
            from mysql_service import is_mysql_running
            if is_mysql_running():
                return True
            print("检测到 MySQL 未运行，正在自动启动...")
            success = start_mysql()
            if success:
                print("[INFO] MySQL 自动启动成功")
                return True
            print("[ERR] MySQL 自动启动失败，请手动初始化数据目录")
            return False
        except Exception as e:
            print(f" 检查/启动 MySQL 时出错: {e}")
            return False

    def _InitDatabase(self) -> None:
        """初始化数据库连接和表结构。"""
        Config = _GetDefaultDbConfig()
        self._Db = MySqlManager(Config)
        self._Db.EnsureSchema()

    def _DbConfigDialog(self, Message: str = "") -> Optional[DatabaseConfig]:
        """弹出数据库配置对话框，用户可修改连接参数。返回新配置或 None（取消）。"""
        print(f"[_DbConfigDialog] 创建配置窗口... (msg={Message[:40] if Message else '无'})")
        Dialog = tk.Toplevel(self)
        Dialog.title("数据库配置")
        Dialog.resizable(False, False)
        # 父窗口隐藏时不设置 transient，否则 Windows 可能自动关闭子窗口
        if self.winfo_viewable():
            Dialog.transient(self)
        Dialog.grab_set()
        Dialog.lift()
        Dialog.focus_set()

        Frame = ttk.Frame(Dialog, padding=15)
        Frame.pack(fill=tk.BOTH, expand=True)

        RowIdx = 0
        if Message:
            ttk.Label(Frame, text=Message, foreground="#CC0000", wraplength=380).grid(
                row=0, column=0, columnspan=2, pady=(0, 12))
            RowIdx = 1

        DefaultConfig = _GetDefaultDbConfig()

        ttk.Label(Frame, text="主机：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        HostVar = tk.StringVar(value=DefaultConfig.Host)
        ttk.Entry(Frame, textvariable=HostVar, width=30).grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        ttk.Label(Frame, text="端口：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        PortVar = tk.StringVar(value=str(DefaultConfig.Port))
        ttk.Entry(Frame, textvariable=PortVar, width=30).grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        ttk.Label(Frame, text="用户名：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        UserVar = tk.StringVar(value=DefaultConfig.User)
        ttk.Entry(Frame, textvariable=UserVar, width=30).grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        ttk.Label(Frame, text="密码：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        PasswordVar = tk.StringVar(value=DefaultConfig.Password)
        ttk.Entry(Frame, textvariable=PasswordVar, width=30, show="*").grid(row=RowIdx, column=1, pady=3)
        RowIdx += 1

        ttk.Label(Frame, text="数据库：").grid(row=RowIdx, column=0, sticky=tk.W, pady=3)
        DbVar = tk.StringVar(value=DefaultConfig.Database)
        ttk.Entry(Frame, textvariable=DbVar, width=30).grid(row=RowIdx, column=1, pady=3)

        Result: List[Optional[DatabaseConfig]] = [None]

        def OnOk() -> None:
            try:
                NewConfig = DatabaseConfig(
                    Host=HostVar.get().strip(),
                    Port=int(PortVar.get().strip()),
                    User=UserVar.get().strip(),
                    Password=PasswordVar.get().strip(),
                    Database=DbVar.get().strip(),
                )
                _SaveDbConfig(NewConfig)
                Result[0] = NewConfig
                Dialog.destroy()
            except ValueError:
                messagebox.showwarning("提示", "端口号必须为数字", parent=Dialog)

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.grid(row=RowIdx + 1, column=0, columnspan=2, pady=(12, 0))
        ttk.Button(BtnFrame, text="确定", command=OnOk).pack(side=tk.RIGHT, padx=2)
        ttk.Button(BtnFrame, text="取消", command=Dialog.destroy).pack(side=tk.RIGHT)

        _CenterOnScreen(Dialog, self)
        Dialog.wait_window()
        return Result[0]

    def _StartDbHealthCheck(self) -> None:
        """启动数据库连接健康检查（每60秒）。"""
        self.after(60000, self._DbHealthCheck)

    def _DbHealthCheck(self) -> None:
        """每60秒检查数据库连接，断联时弹出警告。"""
        if self._Db is None:
            return
        try:
            # 执行简单查询验证连接是否有效
            self._Db.EnsureConnection()
            self._Db.ExecuteQuery("SELECT 1")
        except Exception:
            messagebox.showwarning("连接断开", "数据库连接已断开！")
        # 继续定期检查
        self.after(60000, self._DbHealthCheck)

    def _OnDbConfig(self) -> None:
        """管理员修改数据库连接配置（菜单入口）。"""
        Config = self._DbConfigDialog("可在此修改数据库连接参数，修改后将使用新配置重新连接。")
        if Config is None:
            return
        if self._Db is not None:
            self._Db.Close()
        self._Db = MySqlManager(Config)
        try:
            self._Db.EnsureSchema()
            messagebox.showinfo("成功", "数据库配置已更新")
        except mysql.connector.Error as e:
            messagebox.showerror("错误", f"连接失败：{e}\n请重新配置。")

    def _BuildMainWindow(self) -> None:
        """构建主窗口的 Tab 页面。"""
        self._Notebook = ttk.Notebook(self)
        self._Notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 根据角色动态显示 Tab 页
        Pages: List[Tuple[type[BasePage], str]] = []
        
        if self._User.Role == ROLE_ADMIN or self._User.Role == ROLE_STAFF:
            # 管理员和普通老师可以访问所有管理功能
            Pages.extend([
                (DepartmentPage, "部门管理"),
                (MemberPage, "成员管理"),
                (AttendancePage, "考勤管理"),
                (UserPage, "用户管理"),
                (StatsPage, "统计图表"),
            ])
        elif self._User.Role == ROLE_STUDENT:
            # 学生只能查看自己的考勤记录和统计图表
            Pages.extend([
                (AttendancePage, "我的考勤"),
                (StatsPage, "统计图表"),
            ])

        for PageClass, Title in Pages:
            Page = PageClass(self._Notebook, self._Db, self._User)
            self._Notebook.add(Page, text=Title)

        # 菜单栏
        MenuBar = tk.Menu(self)
        self.config(menu=MenuBar)

        FileMenu = tk.Menu(MenuBar, tearoff=0)
        MenuBar.add_cascade(label="选项", menu=FileMenu)
        FileMenu.add_command(label="重新登录", command=self._ReLogin)
        FileMenu.add_command(label="修改密码", command=self._ChangePassword)
        FileMenu.add_separator()
        FileMenu.add_command(label="退出", command=self.destroy)

        if self._User is not None and self._User.Role == ROLE_ADMIN:
            DbMenu = tk.Menu(MenuBar, tearoff=0)
            MenuBar.add_cascade(label="数据库服务", menu=DbMenu)
            DbMenu.add_command(label="数据库配置", command=self._OnDbConfig)
            DbMenu.add_separator()
            if start_mysql is not None:
                DbMenu.add_command(label="启动 MySQL", command=start_mysql)
                DbMenu.add_command(label="停止 MySQL", command=stop_mysql)
                DbMenu.add_command(label="初始化数据目录", command=initialize_mysql)

        HelpMenu = tk.Menu(MenuBar, tearoff=0)
        MenuBar.add_cascade(label="帮助", menu=HelpMenu)
        HelpMenu.add_command(label="关于", command=self._ShowAbout)

    def _ReLogin(self) -> None:
        """重新登录：隐藏主窗口 → 销毁页面 → 弹出登录框。"""
        if self._Db is None:
            return
        # 隐藏主窗口并销毁当前页面
        self.withdraw()
        if self._Notebook is not None:
            self._Notebook.destroy()
            self._Notebook = None
        # 弹出登录框
        LoginDlg = LoginDialog(self, self._Db)
        if LoginDlg.Result is None:
            # 取消登录则关闭整个程序
            self.destroy()
            return
        self._User = LoginDlg.Result
        self.deiconify()
        self.title(f"实验室考勤管理系统 - {self._User.DisplayName} ({self._User.Role})")
        # 重建所有页面
        self._BuildMainWindow()

    def _ChangePassword(self) -> None:
        """修改当前登录用户的密码。"""
        Dialog = tk.Toplevel(self)
        Dialog.title("修改密码")
        Dialog.resizable(False, False)
        Dialog.grab_set()
        Dialog.transient(self)

        Frame = ttk.Frame(Dialog, padding=15)
        Frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(Frame, text="当前密码：").grid(row=0, column=0, sticky=tk.W, pady=5)
        OldVar = tk.StringVar()
        ttk.Entry(Frame, textvariable=OldVar, width=25, show="*").grid(row=0, column=1, pady=5)

        ttk.Label(Frame, text="新密码：").grid(row=1, column=0, sticky=tk.W, pady=5)
        NewVar = tk.StringVar()
        ttk.Entry(Frame, textvariable=NewVar, width=25, show="*").grid(row=1, column=1, pady=5)

        ttk.Label(Frame, text="确认新密码：").grid(row=2, column=0, sticky=tk.W, pady=5)
        ConfirmVar = tk.StringVar()
        ttk.Entry(Frame, textvariable=ConfirmVar, width=25, show="*").grid(row=2, column=1, pady=5)

        def OnOk() -> None:
            Old = OldVar.get()
            New = NewVar.get()
            Confirm = ConfirmVar.get()
            if not Old or not New:
                messagebox.showwarning("提示", "请填写完整", parent=Dialog)
                return
            if New != Confirm:
                messagebox.showwarning("提示", "两次输入的新密码不一致", parent=Dialog)
                return
            if self._Db is None or self._User is None:
                return
            Success = self._Db.ChangePassword(self._User.UserId, Old, New)
            if Success:
                messagebox.showinfo("成功", "密码修改成功", parent=Dialog)
                Dialog.destroy()
            else:
                messagebox.showerror("错误", "当前密码不正确", parent=Dialog)

        BtnFrame = ttk.Frame(Frame)
        BtnFrame.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(BtnFrame, text="确定", command=OnOk).pack(side=tk.RIGHT, padx=2)
        ttk.Button(BtnFrame, text="取消", command=Dialog.destroy).pack(side=tk.RIGHT)

        Dialog.bind("<Return>", lambda e: OnOk())
        _CenterOnScreen(Dialog, self)
        Dialog.wait_window()

    @staticmethod
    def _ShowAbout() -> None:
        messagebox.showinfo("关于", "实验室考勤与权限管理系统\nVersion 1.0")
